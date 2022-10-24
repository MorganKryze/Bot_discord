import discord
from discord.ext import commands
import youtube_dl
import asyncio
from openpyxl import load_workbook

ytdl = youtube_dl.YoutubeDL()
musics = {}

def setup(bot):
    bot.add_cog(cogAudio(bot))

def searchPlaylist(name,wbplay):
    wsplayA=wbplay[f"{name}playlist"]
    nbLigne=wsplayA.max_row
    PlaylistNames=[]
    for i in range(1,nbLigne+1):
        if wsplayA[f"A{i}"].value !=None:
            PlaylistNames+=[wsplayA[f"A{i}"].value]
    return(PlaylistNames)

def searchPlaylistList(name,plname,wbplay):
    wsplayA=wbplay[f"{name}playlist"]
    position=searchPlaylist(name,wbplay).index(plname)
    i=66
    listePlaylist=[]
    while i>0:
        if  wsplayA[f"{chr(i)}{position+1}"].value!= None and wsplayA[f"{chr(i)}{position+1}"].value!= " ":
            listePlaylist+=[wsplayA[f"{chr(i)}{position+1}"].value]
            i+=1
        else:
            i=0
    return(listePlaylist)

def loadPlaylistExcel():
    wbplay= load_workbook('playlist.xlsx')
    wsplay=wbplay["NomId"]
    wbplay.save('playlist.xlsx')
    return(wbplay,wsplay)

class cogAudio(commands.Cog):
    def __init__(self,bot):
        self.bot = bot

    (wbplay,wsplay)=loadPlaylistExcel()
    def loadPlaylistExcel(self):
        wbplay= load_workbook('playlist.xlsx')
        wsplay=wbplay["NomId"]
        wbplay.save('playlist.xlsx')
        return(wbplay,wsplay)
    

    class Video:
        def __init__(self, link):
            if link[0:4]=="http"or link[0:3]=="wwww":
                video = ytdl.extract_info(link, download=False)
            else:
                video = ytdl.extract_info("ytsearch:%s" % link, download=False)['entries'][0]
            video_format = video["formats"][0]
            self.url = video["webpage_url"]
            self.stream_url = video_format["url"]


    def play_song(self,client, queue, song):
        source = discord.PCMVolumeTransformer(discord.FFmpegPCMAudio(song.stream_url
            , before_options = "-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5"))
        
        def next (_):
            if len(queue) > 0:
                new_song = queue[0]
                del queue[0]
                self.play_song(client, queue, new_song)
            else:
                asyncio.run_coroutine_threadsafe(client.disconnect(), self.bot.loop)

        client.play(source, after = next)
        

    #
    @commands.command()
    async def skip(self,ctx):
        client = ctx.guild.voice_client
        client.stop()
    #    play_song(client, musics[ctx.guild], musics[ctx.guild][0])

    @commands.command()
    async def leave(self,ctx):
        client = ctx.guild.voice_client
        await client.disconnect()
        musics[ctx.guild] = []

    @commands.command()
    async def resume(self,ctx):
        client = ctx.guild.voice_client
        if client.is_paused():
            client.resume()

    @commands.command()
    async def pause(self,ctx):
        client = ctx.guild.voice_client
        if not client.is_paused():
            client.pause()

    @commands.command(aliases= ['p'])
    async def play(self,ctx, *url):
        print("play")
        urla = str(" ".join(url))
        client = ctx.guild.voice_client

        if client and client.channel:
            video = self.Video(urla)
            musics[ctx.guild].append(video)
#            embed= discord.Embed(
#                title="La musique suivante à été ajoutée à la queue:",
#                description=f"{video['title']}"
#            )
#            await ctx.send (embed=embed,delete_after=10)   
        elif ctx.author.voice!=None:
            channel = ctx.author.voice.channel
            video = self.Video(urla)
            musics[ctx.guild]=[]
            client = await channel.connect()
            await ctx.send(f"Je lance : {video.url}",delete_after = 15)
            await ctx.message.delete()
            self.play_song(client, musics[ctx.guild], video)
        else:
            embed= discord.Embed(
                title=f"Utilisateur non connecté",
                description=f"{ctx.author.mention} Veuillez vous connecter à un salon vocal \n||Cette requête s'arrêtera dans 10 secondes||"
            )
            sent=await ctx.send (embed=embed,delete_after=10)        

    @commands.command(aliases= ['PlaylistPlay','plyalistPlay','pPlay','pplay'])
    async def PPlay(self,ctx, *playlist):
        playlist = str(" ".join(playlist))
        client = ctx.guild.voice_client
        (wbplay,wsplay)=loadPlaylistExcel()
        listePlaylist=searchPlaylistList(ctx.message.author.name,playlist,wbplay)
        print(listePlaylist)
        if client and client.channel:
            print(1)
            for song in listePlaylist:
                video = self.Video(song)
                musics[ctx.guild].append(video)
        elif ctx.author.voice!=None:
            print(2)
            channel = ctx.author.voice.channel
            video = self.Video(listePlaylist[0])
            musics[ctx.guild]=[]
            client = await channel.connect()
            await ctx.send(f"Je lance : {video.url}",delete_after = 15)
            await ctx.message.delete()
            self.play_song(client, musics[ctx.guild], video)
            del listePlaylist[0]
            for song in listePlaylist:
                video = self.Video(song)
                musics[ctx.guild].append(video)
        else:
            embed= discord.Embed(
                title=f"Utilisateur non connecté",
                description=f"{ctx.author.mention} Veuillez vous connecter à un salon vocal \n||Cette requête s'arrêtera dans 10 secondes||"
            )
            sent=await ctx.send (embed=embed,delete_after=10)    

